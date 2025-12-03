from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.http import HttpResponse, FileResponse, Http404
import csv, os

# ==================================
# START: IMPORT THƯ VIỆN MỚI
# ==================================
import openpyxl # Thư viện xử lý Excel
from openpyxl.utils import get_column_letter
from io import BytesIO # Thư viện để lưu file vào bộ nhớ
# ==================================
# END: IMPORT
# ==================================

from .models import Task, Evidence, Evaluation
from .forms import TaskForm, EvidenceForm, EvaluationForm


# ---------- Helpers ----------
def is_head(user):
    return user.groups.filter(name__iexact='Head').exists()


# ---------- Auth ----------
def login_view(request):
    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '').strip()
        user = authenticate(request, username=username, password=password)
        if user:
            login(request, user)
            return redirect('dashboard')
        messages.error(request, 'Sai tên đăng nhập hoặc mật khẩu.')
    return render(request, 'workflow/login.html')


def logout_view(request):
    logout(request)
    return redirect('login')


# ---------- Dashboard ----------
@login_required
def dashboard(request):
    user = request.user
    total = Task.objects.count()
    doing = Task.objects.filter(status='DOING').count()
    done = Task.objects.filter(status='DONE').count()
    my_tasks = Task.objects.filter(assignee=user).count()
    is_head_user = is_head(user)
    
    if user.is_superuser or is_head_user:
        tasks = Task.objects.order_by('-id')[:8]
    else:
        tasks = Task.objects.filter(assignee=user).order_by('-id')[:8]

    return render(request, 'workflow/dashboard.html', {
        'total': total, 'doing': doing, 'done': done,
        'my_tasks': my_tasks, 'tasks': tasks,
        'is_head_user': is_head_user
    })


# ---------- Task ----------
@login_required
def task_list(request):
    if request.user.is_superuser or is_head(request.user):
        tasks = Task.objects.all().order_by('-id')
    else:
        tasks = Task.objects.filter(assignee=request.user).order_by('-id')
    return render(request, 'workflow/task_list.html', {'tasks': tasks})


@login_required
def task_detail(request, task_id):
    task = get_object_or_404(Task, pk=task_id)
    evidences = Evidence.objects.filter(task=task)
    evaluations = Evaluation.objects.filter(task=task).order_by('-created_at')

    can_submit = (request.user == task.assignee)
    can_approve = (is_head(request.user) and task.status != 'APPROVED')
    can_evaluate = (is_head(request.user) or request.user.is_superuser)
    is_head_user = is_head(request.user)

    return render(request, 'workflow/task_detail.html', {
        'task': task,
        'evidences': evidences,
        'evaluations': evaluations,
        'can_submit': can_submit,
        'can_approve': can_approve,
        'can_evaluate': can_evaluate,
        'is_head_user': is_head_user,
    })


@login_required
def task_create(request):
    if not (request.user.is_superuser or is_head(request.user)):
        messages.error(request, "Bạn không có quyền tạo công việc.")
        return redirect('task_list')

    if request.method == 'POST':
        form = TaskForm(request.POST)
        if form.is_valid():
            task = form.save(commit=False)
            task.created_by = request.user
            task.status = 'NEW'
            task.save()
            messages.success(request, "Đã tạo công việc mới.")
            return redirect('task_list')
    else:
        form = TaskForm()
    return render(request, 'workflow/task_form.html', {'form': form})


@login_required
def task_edit(request, task_id):
    task = get_object_or_404(Task, pk=task_id)
    if not (request.user.is_superuser or is_head(request.user)):
        messages.error(request, "Bạn không có quyền sửa công việc.")
        return redirect('task_list')

    if request.method == 'POST':
        form = TaskForm(request.POST, instance=task)
        if form.is_valid():
            form.save()
            messages.success(request, "Đã cập nhật công việc.")
            return redirect('task_detail', task_id=task.id)
    else:
        form = TaskForm(instance=task)
    return render(request, 'workflow/task_form.html', {'form': form, 'edit': True})


@login_required
def task_delete(request, task_id):
    if not (request.user.is_superuser or is_head(request.user)):
        messages.error(request, "Bạn không có quyền xóa.")
        return redirect('task_list')
    task = get_object_or_404(Task, pk=task_id)
    task.delete()
    messages.success(request, "Đã xóa công việc.")
    return redirect('task_list')


# ---------- Evidence ----------
@login_required
def submit_evidence(request, task_id):
    task = get_object_or_404(Task, pk=task_id)
    if request.user != task.assignee:
        messages.error(request, "Bạn chỉ được nộp minh chứng cho công việc của mình.")
        return redirect('task_detail', task_id=task.id)

    if request.method == 'POST':
        form = EvidenceForm(request.POST, request.FILES)
        if form.is_valid():
            ev = form.save(commit=False)
            ev.task = task
            ev.uploader = request.user
            ev.save()
            messages.success(request, "Đã nộp minh chứng.")
            return redirect('task_detail', task_id=task.id)
    else:
        form = EvidenceForm()
    return render(request, 'workflow/submit_form.html', {'form': form, 'task': task})


@login_required
def delete_evidence(request, ev_id):
    ev = get_object_or_404(Evidence, pk=ev_id)
    if request.user != ev.uploader:
        messages.error(request, "Bạn không được phép xóa minh chứng này.")
        return redirect('task_detail', task_id=ev.task.id)
    
    ev.file.delete(save=False)
    ev.delete()
    messages.success(request, "Đã hủy nộp minh chứng.")
    return redirect('task_detail', task_id=ev.task.id)


@login_required
def download_evidence(request, ev_id):
    ev = get_object_or_404(Evidence, pk=ev_id)
    file_path = ev.file.path
    if not os.path.exists(file_path):
        raise Http404("Không tìm thấy file.")
    return FileResponse(open(file_path, 'rb'), as_attachment=True, filename=os.path.basename(file_path))


# ---------- Evaluation ----------
@login_required
def evaluate_task(request, task_id):
    if not (is_head(request.user) or request.user.is_superuser):
        messages.error(request, "Bạn không có quyền đánh giá.")
        return redirect('task_list')

    task = get_object_or_404(Task, pk=task_id)
    if request.method == 'POST':
        form = EvaluationForm(request.POST)
        if form.is_valid():
            e = form.save(commit=False)
            e.task = task
            e.evaluator = request.user
            e.save()
            task.status = 'EVAL'
            task.save()
            messages.success(request, "Đã đánh giá công việc.")
            return redirect('task_detail', task_id=task.id)
    else:
        form = EvaluationForm()
    return render(request, 'workflow/evaluate_form.html', {'form': form, 'task': task})


@login_required
def approve_task(request, task_id):
    if not (request.user.is_superuser or is_head(request.user)):
        messages.error(request, "Bạn không có quyền phê duyệt.")
        return redirect('task_list')

    task = get_object_or_404(Task, pk=task_id)
    task.status = 'APPROVED'
    task.save()
    messages.success(request, "✅ Đã phê duyệt hoàn thành công việc.")
    return redirect('task_detail', task_id=task.id)


# ---------- Update Status (Trưởng bộ môn) ----------
@login_required
def update_status(request, task_id):
    """Hiển thị form dropdown chọn trạng thái mới"""
    if not (request.user.is_superuser or is_head(request.user)):
        messages.error(request, "Bạn không có quyền cập nhật trạng thái.")
        return redirect('task_detail', task_id=task_id)

    task = get_object_or_404(Task, pk=task_id)

    if request.method == 'POST':
        new_status = request.POST.get('status')
        if new_status in dict(Task.STATUS_CHOICES).keys():
            task.status = new_status
            task.save()
            messages.success(request, f"✅ Đã cập nhật trạng thái thành: {task.get_status_display()}")
        else:
            messages.error(request, "Trạng thái không hợp lệ.")
        return redirect('task_detail', task_id=task.id)

    return render(request, 'workflow/update_status.html', {
        'task': task,
        'status_choices': Task.STATUS_CHOICES,
    })



# ---------- Reports ----------
@login_required
def reports(request):
    all_status = dict(Task.STATUS_CHOICES)
    by_status = [(v, Task.objects.filter(status=k).count()) for k, v in all_status.items()]
    total = Task.objects.count()
    return render(request, 'workflow/reports.html', {
        'by_status': by_status, 'total': total
    })


# ==================================
# START: THAY THẾ HÀM XUẤT FILE
# ==================================
@login_required
def export_tasks_excel(request): # Đổi tên hàm
    if not (request.user.is_superuser or is_head(request.user)):
        messages.error(request, "Bạn không có quyền xuất báo cáo.")
        return redirect('dashboard')

    # Tạo một workbook Excel mới
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bao Cao Cong Viec"

    # Định nghĩa tiêu đề cột
    columns = ['Tên công việc', 'Loại', 'Ưu tiên', 'Trạng thái', 'Bắt đầu', 'Hạn chót', 'Người phụ trách']
    ws.append(columns)

    # Lấy dữ liệu từ database và ghi vào file
    for t in Task.objects.all().order_by('-id'):
        ws.append([
            t.title,
            dict(Task.TYPE_CHOICES).get(t.type),
            dict(Task.PRIORITY_CHOICES).get(t.priority),
            dict(Task.STATUS_CHOICES).get(t.status),
            t.start_date,
            t.due_date,
            t.assignee.username if t.assignee else '',
        ])

    # Tự động điều chỉnh độ rộng cột
    for i, column_cells in enumerate(ws.columns):
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        adjusted_width = max_length + 2
        ws.column_dimensions[get_column_letter(i + 1)].width = adjusted_width

    # Lưu file vào bộ nhớ
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Trả file về cho người dùng
    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="bao_cao_cong_viec.xlsx"' # Đổi đuôi file
    return response

# ==================================
# END: THAY THẾ HÀM XUẤT FILE
# ==================================